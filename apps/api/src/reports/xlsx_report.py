"""Relatório mensal em Excel — mesmos dados do PDF, mas em células, para
quem quer continuar a conta na planilha em vez de só ler o resultado.

openpyxl (Python puro, sem dependência de sistema) pelo mesmo motivo que o
PDF usa fpdf2: a função da Vercel tem teto de bundle e não comporta as
libs pesadas de planilha.
"""
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_NAVY = "0A192F"
_HEADER_FILL = PatternFill("solid", start_color=_NAVY, end_color=_NAVY)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(bold=True, size=14, color=_NAVY)
_SECTION_FONT = Font(bold=True, size=11, color=_NAVY)
_MUTED_FONT = Font(italic=True, size=9, color="64748B")

_BRL = 'R$ #,##0.00'
_PCT = '0.0%'

_MONTH_NAMES_PT = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

# Espelha ASSET_TYPE_LABELS do frontend (chartTheme.ts).
_ASSET_TYPE_LABELS = {
    "stock": "Ações",
    "stock_br": "Ações BR",
    "stock_us": "Ações EUA",
    "fii": "FIIs",
    "reit": "REITs",
    "etf": "ETFs",
    "crypto": "Cripto",
    "commodity": "Commodities",
    "fixed_income_br": "Renda Fixa",
    "other": "Outros",
}


def _num(value) -> float:
    """openpyxl não serializa Decimal — e gravar texto formatado perderia a
    capacidade de somar na planilha, que é o motivo de existir o Excel."""
    return float(value if value is not None else 0)


def _write_header(ws, row: int, labels: list[str]) -> int:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    return row + 1


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 4, 12), 48)


def _finance_sheet(wb: Workbook, title: str, summary: dict, *, include_charts: bool = True) -> None:
    ws = wb.create_sheet(title=title[:31])
    ws["A1"] = title
    ws["A1"].font = _SECTION_FONT

    row = 3
    for label, key in (("Receitas", "income"), ("Despesas", "expense"), ("Saldo", "net")):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        cell = ws.cell(row=row, column=2, value=_num(summary.get(key)))
        cell.number_format = _BRL
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Despesas por categoria").font = _SECTION_FONT
    row += 1
    by_category = summary.get("by_category") or []
    if by_category:
        cat_header_row = row
        row = _write_header(ws, row, ["Categoria", "Valor", "% do mês"])
        for cat in by_category:
            ws.cell(row=row, column=1, value=cat["category_name"])
            ws.cell(row=row, column=2, value=_num(cat["value"])).number_format = _BRL
            ws.cell(row=row, column=3, value=_num(cat["pct"])).number_format = _PCT
            row += 1
        if include_charts:
            # Pizza é legítima aqui: a soma das categorias é o total de
            # despesas, então cada fatia é fração de um todo real.
            pie = PieChart()
            pie.title = "Despesas por categoria"
            pie.height, pie.width = 8, 13
            pie.add_data(Reference(ws, min_col=2, min_row=cat_header_row, max_row=row - 1), titles_from_data=True)
            pie.set_categories(Reference(ws, min_col=1, min_row=cat_header_row + 1, max_row=row - 1))
            ws.add_chart(pie, "F3")
    else:
        ws.cell(row=row, column=1, value="Nenhuma despesa registrada neste mês.").font = _MUTED_FONT
        row += 1

    # Série mensal — a base dos gráficos de evolução. Fica em células (não só
    # no gráfico) para quem quiser refazer a conta na planilha.
    monthly = summary.get("monthly_series") or []
    if monthly:
        row += 2
        ws.cell(row=row, column=1, value="Evolução mensal").font = _SECTION_FONT
        row += 1
        series_header = row
        row = _write_header(ws, row, ["Mês", "Receitas", "Despesas", "Saldo"])
        for point in monthly:
            income, expense = _num(point.get("income")), _num(point.get("expense"))
            ws.cell(row=row, column=1, value=point.get("month"))
            ws.cell(row=row, column=2, value=income).number_format = _BRL
            ws.cell(row=row, column=3, value=expense).number_format = _BRL
            ws.cell(row=row, column=4, value=income - expense).number_format = _BRL
            row += 1

        if include_charts and len(monthly) >= 2:
            bars = BarChart()
            bars.type, bars.grouping = "col", "clustered"
            bars.title = "Receitas x despesas"
            bars.y_axis.title, bars.x_axis.title = "R$", "Mês"
            bars.height, bars.width = 8, 18
            bars.add_data(
                Reference(ws, min_col=2, max_col=3, min_row=series_header, max_row=row - 1),
                titles_from_data=True,
            )
            bars.set_categories(Reference(ws, min_col=1, min_row=series_header + 1, max_row=row - 1))
            ws.add_chart(bars, f"F{series_header}")

            line = LineChart()
            line.title = "Saldo mensal"
            line.y_axis.title, line.x_axis.title = "R$", "Mês"
            line.height, line.width = 8, 18
            line.add_data(
                Reference(ws, min_col=4, min_row=series_header, max_row=row - 1),
                titles_from_data=True,
            )
            line.set_categories(Reference(ws, min_col=1, min_row=series_header + 1, max_row=row - 1))
            ws.add_chart(line, f"F{series_header + 17}")

    _autosize(ws)


def _portfolios_sheet(wb: Workbook, portfolios: list[dict], *, include_charts: bool = True) -> None:
    ws = wb.create_sheet(title="Investimentos")
    ws["A1"] = "Investimentos"
    ws["A1"].font = _SECTION_FONT

    row = 3
    if not portfolios:
        ws.cell(row=row, column=1, value="Nenhuma carteira de investimentos selecionada.").font = _MUTED_FONT
        _autosize(ws)
        return

    row = _write_header(ws, row, ["Carteira", "Investido", "Valor atual", "P&L", "P&L %"])
    first_data_row = row
    for p in portfolios:
        ws.cell(row=row, column=1, value=p["portfolio_name"])
        ws.cell(row=row, column=2, value=_num(p["total_invested_brl"])).number_format = _BRL
        ws.cell(row=row, column=3, value=_num(p["total_market_value_brl"])).number_format = _BRL
        ws.cell(row=row, column=4, value=_num(p["total_pnl_absolute"])).number_format = _BRL
        # total_pnl_percent vem em pontos percentuais; o formato de célula do
        # Excel multiplica por 100 de novo, daí a divisão.
        ws.cell(row=row, column=5, value=_num(p["total_pnl_percent"]) / 100).number_format = _PCT
        row += 1

    # Fórmula em vez de valor calculado: quem abrir a planilha e apagar uma
    # linha vê o total se corrigir sozinho, que é o ponto de exportar Excel.
    last = row - 1
    ws.cell(row=row, column=1, value="Total").font = Font(bold=True, size=10)
    for col in (2, 3, 4):
        letter = get_column_letter(col)
        cell = ws.cell(row=row, column=col, value=f"=SUM({letter}{first_data_row}:{letter}{last})")
        cell.number_format = _BRL
        cell.font = Font(bold=True, size=10)
    row += 2

    # Alocação somada entre as carteiras — a exposição real por classe, que
    # nenhuma carteira isolada mostra.
    allocation: dict[str, float] = {}
    for p in portfolios:
        for entry in p.get("allocation_by_type") or []:
            name = _ASSET_TYPE_LABELS.get(entry.get("asset_type") or "other", entry.get("asset_type") or "Outros")
            allocation[name] = allocation.get(name, 0.0) + _num(entry.get("value"))

    if allocation:
        ws.cell(row=row, column=1, value="Alocação por classe de ativo").font = _SECTION_FONT
        row += 1
        alloc_header = row
        row = _write_header(ws, row, ["Classe", "Valor"])
        for name, value in sorted(allocation.items(), key=lambda kv: kv[1], reverse=True):
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=value).number_format = _BRL
            row += 1

        if include_charts:
            pie = PieChart()
            pie.title = "Alocação por classe de ativo"
            pie.height, pie.width = 8, 13
            pie.add_data(Reference(ws, min_col=2, min_row=alloc_header, max_row=row - 1), titles_from_data=True)
            pie.set_categories(Reference(ws, min_col=1, min_row=alloc_header + 1, max_row=row - 1))
            ws.add_chart(pie, f"H{alloc_header}")

    if include_charts and len(portfolios) > 1:
        bars = BarChart()
        bars.type = "bar"
        bars.title = "Valor de mercado por carteira"
        bars.height, bars.width = 8, 15
        bars.add_data(
            Reference(ws, min_col=3, min_row=first_data_row - 1, max_row=last),
            titles_from_data=True,
        )
        bars.set_categories(Reference(ws, min_col=1, min_row=first_data_row, max_row=last))
        ws.add_chart(bars, "H3")

    _autosize(ws)


def generate_monthly_report_xlsx(
    *,
    user_name: str,
    month: str,
    finance_sections: list[tuple[str, dict]],
    portfolios: list[dict],
    include_finance: bool = True,
    include_investments: bool = True,
    include_charts: bool = True,
) -> bytes:
    """finance_sections: (rótulo da carteira, resumo) — uma aba por carteira
    selecionada, ou uma única "Consolidado" quando nenhuma foi escolhida.

    include_investments: com False a aba de investimentos não é criada — uma
    carteira pode ser de outra pessoa e não deve constar do relatório."""
    year, mon = int(month[:4]), int(month[5:7])
    period_label = f"{_MONTH_NAMES_PT[mon]} de {year}"

    wb = Workbook()
    cover = wb.active
    cover.title = "Resumo"
    cover["A1"] = "InvestIQ - Relatório Mensal"
    cover["A1"].font = _TITLE_FONT
    cover["A3"] = "Usuário"
    cover["B3"] = user_name
    cover["A4"] = "Período"
    cover["B4"] = period_label
    cover["A5"] = "Gerado em"
    cover["B5"] = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    cover["A7"] = (
        "Cotações via fontes públicas gratuitas. "
        "Este relatório não constitui recomendação de investimento."
    )
    cover["A7"].font = _MUTED_FONT
    for row in range(3, 6):
        cover.cell(row=row, column=1).font = Font(bold=True, size=10)
    _autosize(cover)

    if include_finance:
        for label, summary in finance_sections:
            _finance_sheet(wb, label, summary, include_charts=include_charts)
    if include_investments:
        _portfolios_sheet(wb, portfolios, include_charts=include_charts)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
