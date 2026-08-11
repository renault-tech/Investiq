"""Relatório mensal em Excel — uma aba por carteira selecionada, com
valores como número (e não texto formatado), que é o motivo de exportar
planilha em vez de PDF."""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from src.reports.xlsx_report import generate_monthly_report_xlsx


def _summary(income: str, expense: str, categories=None) -> dict:
    return {
        "income": Decimal(income),
        "expense": Decimal(expense),
        "net": Decimal(income) - Decimal(expense),
        "by_category": categories or [],
    }


def _load(content: bytes):
    return load_workbook(BytesIO(content))


def test_uma_aba_por_carteira_mais_o_resumo_e_investimentos():
    content = generate_monthly_report_xlsx(
        user_name="Teste", month="2026-07",
        finance_sections=[("Nubank", _summary("5000", "3000")), ("Itau", _summary("2000", "800"))],
        portfolios=[],
    )
    wb = _load(content)
    assert wb.sheetnames == ["Resumo", "Nubank", "Itau", "Investimentos"]


def test_valores_saem_como_numero_para_a_planilha_poder_somar():
    content = generate_monthly_report_xlsx(
        user_name="Teste", month="2026-07",
        finance_sections=[("Consolidado", _summary("5000.50", "3000.25"))],
        portfolios=[],
    )
    ws = _load(content)["Consolidado"]
    assert ws["B3"].value == 5000.50
    assert ws["B4"].value == 3000.25
    assert ws["B5"].value == 2000.25
    assert ws["B3"].number_format == "R$ #,##0.00"


def test_categorias_viram_linhas_com_percentual():
    content = generate_monthly_report_xlsx(
        user_name="Teste", month="2026-07",
        finance_sections=[(
            "Consolidado",
            _summary("5000", "3200", [
                {"category_name": "Moradia", "value": Decimal("1200.00"), "pct": Decimal("0.375")},
            ]),
        )],
        portfolios=[],
    )
    ws = _load(content)["Consolidado"]
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    moradia = next(r for r in rows if r and r[0] == "Moradia")
    assert moradia[1] == 1200.00
    assert moradia[2] == 0.375


def test_total_de_investimentos_sai_como_formula():
    """Fórmula e não valor: quem apagar uma linha na planilha vê o total se
    corrigir sozinho, que é o ponto de exportar Excel."""
    content = generate_monthly_report_xlsx(
        user_name="Teste", month="2026-07",
        finance_sections=[("Consolidado", _summary("0", "0"))],
        portfolios=[
            {
                "portfolio_name": "Principal",
                "total_invested_brl": Decimal("10000.00"),
                "total_market_value_brl": Decimal("11500.00"),
                "total_pnl_absolute": Decimal("1500.00"),
                "total_pnl_percent": Decimal("15.00"),
            }
        ],
    )
    ws = _load(content)["Investimentos"]
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    total = next(r for r in rows if r and r[0] == "Total")
    assert str(total[1]).startswith("=SUM(")


def test_pnl_percent_convertido_para_fracao():
    """total_pnl_percent chega em pontos percentuais e o formato de célula do
    Excel multiplica por 100 de novo — sem dividir, 15% viraria 1500%."""
    content = generate_monthly_report_xlsx(
        user_name="Teste", month="2026-07",
        finance_sections=[("Consolidado", _summary("0", "0"))],
        portfolios=[
            {
                "portfolio_name": "Principal",
                "total_invested_brl": Decimal("10000.00"),
                "total_market_value_brl": Decimal("11500.00"),
                "total_pnl_absolute": Decimal("1500.00"),
                "total_pnl_percent": Decimal("15.00"),
            }
        ],
    )
    ws = _load(content)["Investimentos"]
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    principal = next(r for r in rows if r and r[0] == "Principal")
    assert principal[4] == 0.15
