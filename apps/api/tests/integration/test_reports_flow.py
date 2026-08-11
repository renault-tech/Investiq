"""Integration: GET /reports/monthly end-to-end against a real Postgres —
seeds a finance transaction and a portfolio position, then checks the
response is a real PDF with the right headers.

No live market data network in this sandbox -> seed last_price directly on
the asset, same fallback get_portfolio_summary uses when a live quote isn't
available (established pattern, see test_fx_conversion.py).
"""
import pytest
from sqlalchemy import select

from src.portfolio.models import Asset
from .conftest import register_and_login


@pytest.mark.asyncio
async def test_monthly_report_is_a_valid_pdf(client, db_session):
    session = await register_and_login(client)
    headers = session["headers"]

    categories = (await client.get("/finance/categories", headers=headers)).json()
    category_id = next(c["id"] for c in categories if c["category_type"] == "expense")
    await client.post(
        "/finance/transactions",
        json={
            "transaction_type": "expense", "amount": 150, "category_id": category_id,
            "transaction_date": "2026-07-05T12:00:00Z",
        },
        headers=headers,
    )

    portfolio = await client.post("/portfolios/", json={"name": "Principal", "currency": "BRL"}, headers=headers)
    portfolio_id = portfolio.json()["id"]
    position = await client.post(f"/portfolios/{portfolio_id}/positions", json={"ticker": "VALE3"}, headers=headers)
    position_id = position.json()["id"]
    await client.post(
        "/portfolios/transactions",
        json={
            "position_id": position_id, "transaction_type": "buy",
            "quantity": 10, "unit_price": 50, "fees": 0, "fx_rate": 1,
            "transaction_date": "2026-07-01T12:00:00Z",
        },
        headers=headers,
    )
    asset = (await db_session.execute(select(Asset).where(Asset.ticker == "VALE3"))).scalar_one()
    asset.last_price = 55
    await db_session.commit()

    resp = await client.get("/reports/monthly", params={"month": "2026-07"}, headers=headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert "relatorio_2026-07.pdf" in resp.headers["content-disposition"]
    assert resp.content[:5] == b"%PDF-"
    assert len(resp.content) > 1000


@pytest.mark.asyncio
async def test_monthly_report_with_no_data_still_returns_a_pdf(client):
    session = await register_and_login(client)
    resp = await client.get("/reports/monthly", params={"month": "2026-07"}, headers=session["headers"])
    assert resp.status_code == 200
    assert resp.content[:5] == b"%PDF-"


@pytest.mark.asyncio
async def test_monthly_report_rejects_malformed_month(client):
    session = await register_and_login(client)
    resp = await client.get("/reports/monthly", params={"month": "not-a-month"}, headers=session["headers"])
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_monthly_report_requires_auth(client):
    resp = await client.get("/reports/monthly", params={"month": "2026-07"})
    assert resp.status_code in (401, 403)


# --- Excel e seleção de carteiras -----------------------------------------

async def _account(client, headers, name: str) -> str:
    resp = await client.post(
        "/finance/accounts",
        json={"name": name, "holder": "Eu", "account_type": "checking", "opening_balance": 0},
        headers=headers,
    )
    return resp.json()["id"]


async def _expense(client, headers, category_id: str, account_id: str, amount: int) -> None:
    await client.post(
        "/finance/transactions",
        json={
            "transaction_type": "expense", "amount": amount, "category_id": category_id,
            "bank_account_id": account_id, "transaction_date": "2026-07-05T12:00:00Z",
        },
        headers=headers,
    )


@pytest.mark.asyncio
async def test_monthly_report_as_xlsx(client):
    session = await register_and_login(client)
    resp = await client.get(
        "/reports/monthly", params={"month": "2026-07", "format": "xlsx"}, headers=session["headers"]
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "relatorio_2026-07.xlsx" in resp.headers["content-disposition"]
    # xlsx é um zip; PK é a assinatura.
    assert resp.content[:2] == b"PK"


@pytest.mark.asyncio
async def test_monthly_report_rejects_unknown_format(client):
    session = await register_and_login(client)
    resp = await client.get(
        "/reports/monthly", params={"month": "2026-07", "format": "docx"}, headers=session["headers"]
    )
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_xlsx_has_one_sheet_per_selected_account(client):
    from io import BytesIO

    from openpyxl import load_workbook

    session = await register_and_login(client)
    headers = session["headers"]
    categories = (await client.get("/finance/categories", headers=headers)).json()
    category_id = next(c["id"] for c in categories if c["category_type"] == "expense")

    nubank = await _account(client, headers, "Nubank")
    itau = await _account(client, headers, "Itau")
    await _expense(client, headers, category_id, nubank, 200)
    await _expense(client, headers, category_id, itau, 700)

    resp = await client.get(
        "/reports/monthly",
        params={"month": "2026-07", "format": "xlsx", "account_ids": f"{nubank},{itau}"},
        headers=headers,
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    assert "Nubank" in wb.sheetnames
    assert "Itau" in wb.sheetnames
    assert "Consolidado" not in wb.sheetnames


@pytest.mark.asyncio
async def test_sem_selecao_o_relatorio_sai_consolidado(client):
    from io import BytesIO

    from openpyxl import load_workbook

    session = await register_and_login(client)
    resp = await client.get(
        "/reports/monthly", params={"month": "2026-07", "format": "xlsx"}, headers=session["headers"]
    )
    wb = load_workbook(BytesIO(resp.content))
    assert "Consolidado" in wb.sheetnames


@pytest.mark.asyncio
async def test_carteira_de_outro_usuario_e_rejeitada(client):
    a = await register_and_login(client)
    b = await register_and_login(client)
    account_de_a = await _account(client, a["headers"], "Nubank")

    resp = await client.get(
        "/reports/monthly",
        params={"month": "2026-07", "account_ids": account_de_a},
        headers=b["headers"],
    )
    assert resp.status_code == 422, resp.text
